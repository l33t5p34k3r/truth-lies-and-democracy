#!/usr/bin/env python3
import yaml
import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any
import sys

class SchemaMigration:
    def __init__(self, old_schema_path: str, new_schema_path: str):
        with open(old_schema_path, 'r') as f:
            self.old_schema = yaml.safe_load(f)
        with open(new_schema_path, 'r') as f:
            self.new_schema = yaml.safe_load(f)
        
        self.old_types = self.old_schema['types']
        self.new_types = self.new_schema['types']
    
    def analyze_changes(self) -> Dict[str, Any]:
        changes = {
            'added_types': [],
            'removed_types': [],
            'modified_types': {}
        }
        
        old_type_names = set(self.old_types.keys())
        new_type_names = set(self.new_types.keys())
        
        changes['added_types'] = list(new_type_names - old_type_names)
        changes['removed_types'] = list(old_type_names - new_type_names)
        
        for type_name in old_type_names & new_type_names:
            type_changes = self._analyze_type_changes(type_name)
            if type_changes['has_changes']:
                changes['modified_types'][type_name] = type_changes
        
        return changes
    
    def _analyze_type_changes(self, type_name: str) -> Dict[str, Any]:
        old_fields = self.old_types[type_name]['fields']
        new_fields = self.new_types[type_name]['fields']
        
        old_field_names = set(old_fields.keys())
        new_field_names = set(new_fields.keys())
        
        changes = {
            'has_changes': False,
            'added_fields': [],
            'removed_fields': [],
            'type_changes': [],
            'constraint_changes': []
        }
        
        added = list(new_field_names - old_field_names)
        removed = list(old_field_names - new_field_names)
        
        if added:
            changes['added_fields'] = added
            changes['has_changes'] = True
        
        if removed:
            changes['removed_fields'] = removed
            changes['has_changes'] = True
        
        for field_name in old_field_names & new_field_names:
            old_field = old_fields[field_name]
            new_field = new_fields[field_name]
            
            if old_field['type'] != new_field['type']:
                changes['type_changes'].append({ # pyright: ignore[reportAttributeAccessIssue]
                    'field': field_name,
                    'old_type': old_field['type'],
                    'new_type': new_field['type']
                })
                changes['has_changes'] = True
            
            old_constraints = set(str(c) for c in old_field.get('constraints', []))
            new_constraints = set(str(c) for c in new_field.get('constraints', []))
            
            if old_constraints != new_constraints:
                changes['constraint_changes'].append({ # pyright: ignore[reportAttributeAccessIssue]
                    'field': field_name,
                    'added': list(new_constraints - old_constraints),
                    'removed': list(old_constraints - new_constraints)
                })
                changes['has_changes'] = True
        
        return changes
    
    def print_migration_report(self) -> None:
        changes = self.analyze_changes()
        
        print("\n=== SCHEMA MIGRATION REPORT ===\n")
        
        if changes['added_types']:
            print("NEW TYPES:")
            for type_name in changes['added_types']:
                print(f"  + {type_name}")
            print()
        
        if changes['removed_types']:
            print("REMOVED TYPES:")
            for type_name in changes['removed_types']:
                print(f"  - {type_name}")
            print()
        
        if changes['modified_types']:
            print("MODIFIED TYPES:")
            for type_name, type_changes in changes['modified_types'].items():
                print(f"\n  {type_name}:")
                
                if type_changes['added_fields']:
                    print("    Added fields:")
                    for field in type_changes['added_fields']:
                        print(f"      + {field}")
                
                if type_changes['removed_fields']:
                    print("    Removed fields:")
                    for field in type_changes['removed_fields']:
                        print(f"      - {field}")
                
                if type_changes['type_changes']:
                    print("    Type changes:")
                    for change in type_changes['type_changes']:
                        print(f"      ~ {change['field']}: {change['old_type']} → {change['new_type']}")
                
                if type_changes['constraint_changes']:
                    print("    Constraint changes:")
                    for change in type_changes['constraint_changes']:
                        print(f"      ~ {change['field']}")
                        if change['added']:
                            for c in change['added']:
                                print(f"        + {c}")
                        if change['removed']:
                            for c in change['removed']:
                                print(f"        - {c}")
        
        print("\n=== ACTIONS REQUIRED ===\n")
        print("1. Regenerate code: python generate_from_schema.py new_schema.yaml")
        print("2. Migrate existing data files (see migrate_json_data below)")
        print("3. Update Excel templates for editors")
        print("4. Test data loading with new validation rules")
        print()
    
    def migrate_json_data(self, input_path: str, output_path: str) -> None:
        with open(input_path, 'r') as f:
            data = json.load(f)
        
        changes = self.analyze_changes()
        migrated_data = {}
        
        for type_name in data.keys():
            if type_name in changes['removed_types']:
                print(f"Warning: Skipping removed type {type_name}")
                continue
            
            migrated_data[type_name] = []
            
            for entry in data[type_name]:
                migrated_entry = self._migrate_entry(type_name, entry, changes)
                migrated_data[type_name].append(migrated_entry)
        
        for type_name in changes['added_types']:
            migrated_data[type_name] = []
        
        with open(output_path, 'w') as f:
            json.dump(migrated_data, f, indent=2)
        
        print(f"Migrated data saved to {output_path}")
    
    def _migrate_entry(self, type_name: str, entry: Dict, changes: Dict) -> Dict:
        migrated = entry.copy()
        
        if type_name not in changes['modified_types']:
            return migrated
        
        type_changes = changes['modified_types'][type_name]
        
        for field in type_changes['removed_fields']:
            if field in migrated:
                print(f"  Removing deprecated field {type_name}.{field}")
                del migrated[field]
        
        for field in type_changes['added_fields']:
            if field not in migrated:
                default_value = self._get_default_value(type_name, field)
                print(f"  Adding new field {type_name}.{field} = {default_value}")
                migrated[field] = default_value
        
        for change in type_changes['type_changes']:
            field = change['field']
            if field in migrated:
                migrated[field] = self._convert_type(
                    migrated[field],
                    change['old_type'],
                    change['new_type']
                )
        
        return migrated
    
    def _get_default_value(self, type_name: str, field_name: str) -> Any:
        field_def = self.new_types[type_name]['fields'][field_name]
        field_type = field_def['type']
        constraints = field_def.get('constraints', [])
        
        if 'optional' in constraints:
            return None
        
        if field_type == 'int':
            return 0
        elif field_type == 'float':
            return 0.0
        elif field_type == 'string':
            return ""
        elif field_type == 'bool':
            return False
        elif field_type.startswith('array'):
            return []
        else:
            return None
    
    def _convert_type(self, value: Any, old_type: str, new_type: str) -> Any:
        if value is None:
            return None
        
        try:
            if new_type == 'int':
                return int(value)
            elif new_type == 'float':
                return float(value)
            elif new_type == 'string':
                return str(value)
            elif new_type == 'bool':
                if isinstance(value, bool):
                    return value
                return str(value).lower() in ['true', '1', 'yes']
            elif new_type.startswith('array'):
                if isinstance(value, list):
                    return value
                return [value]
        except (ValueError, TypeError) as e:
            print(f"Warning: Failed to convert {value} from {old_type} to {new_type}: {e}")
            return value
        
        return value
    
    def migrate_excel_data(self, old_excel: str, new_excel: str) -> None:
        old_wb = openpyxl.load_workbook(old_excel)
        changes = self.analyze_changes()
        
        from generate_from_schema import SchemaGenerator
        generator = SchemaGenerator.__new__(SchemaGenerator)
        generator.schema = self.new_schema
        generator.types = self.new_types
        generator.generate_excel_template(new_excel) # pyright: ignore[reportArgumentType]
        
        new_wb = openpyxl.load_workbook(new_excel)
        
        for type_name in self.new_types.keys():
            if type_name not in old_wb.sheetnames:
                print(f"New type {type_name} - empty sheet created")
                continue
            
            if type_name in changes['removed_types']:
                print(f"Skipping removed type {type_name}")
                continue
            
            old_ws = old_wb[type_name]
            new_ws = new_wb[type_name]
            
            old_headers = [cell.value for cell in old_ws[1]]
            new_headers = [cell.value for cell in new_ws[1]]
            
            header_mapping = {}
            for i, header in enumerate(old_headers):
                if header in new_headers:
                    new_idx = new_headers.index(header)
                    header_mapping[i] = new_idx
            
            for row_idx, row in enumerate(old_ws.iter_rows(min_row=3, values_only=True), start=3):
                if all(cell is None for cell in row):
                    continue
                
                for old_col_idx, value in enumerate(row):
                    if old_col_idx in header_mapping:
                        new_col_idx = header_mapping[old_col_idx]
                        new_ws.cell(row=row_idx, column=new_col_idx + 1, value=value)
        
        new_wb.save(new_excel)
        print(f"Migrated Excel data saved to {new_excel}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python schema_migration.py <old_schema.yaml> <new_schema.yaml> [command] [args]")
        print("\nCommands:")
        print("  report                                    - Show migration report")
        print("  migrate-json <input.json> <output.json>  - Migrate JSON data")
        print("  migrate-excel <input.xlsx> <output.xlsx> - Migrate Excel data")
        sys.exit(1)
    
    old_schema = sys.argv[1]
    new_schema = sys.argv[2]
    
    migration = SchemaMigration(old_schema, new_schema)
    
    if len(sys.argv) == 3:
        migration.print_migration_report()
    else:
        command = sys.argv[3]
        
        if command == "report":
            migration.print_migration_report()
        
        elif command == "migrate-json":
            if len(sys.argv) != 6:
                print("Usage: ... migrate-json <input.json> <output.json>")
                sys.exit(1)
            migration.migrate_json_data(sys.argv[4], sys.argv[5])
        
        elif command == "migrate-excel":
            if len(sys.argv) != 6:
                print("Usage: ... migrate-excel <input.xlsx> <output.xlsx>")
                sys.exit(1)
            migration.migrate_excel_data(sys.argv[4], sys.argv[5])
        
        else:
            print(f"Unknown command: {command}")
            sys.exit(1)