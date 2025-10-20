#!/usr/bin/env python3
import yaml
import json
from pathlib import Path
from typing import Dict, List, Any
from jinja2 import Environment, FileSystemLoader, Template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class SchemaGenerator:
    def __init__(self, schema_path: str, templates_dir: str = "templates"):
        with open(schema_path, 'r') as f:
            self.schema = yaml.safe_load(f)
        self.types = self.schema['types']
        self.enums = self.schema.get('enums', {})

        self.templates_path = Path(templates_dir)
        self.env = Environment(
            loader=FileSystemLoader(self.templates_path),
            trim_blocks=True,
            lstrip_blocks=True
        )

        self.env.filters['gdscript_type'] = self._map_to_gdscript_type
        self.env.filters['gdscript_default'] = self._get_gdscript_default
        self.env.filters['python_type'] = self._map_to_python_type
        self.env.filters['get_id_field'] = self._get_id_field

        self._analyze_references()

    def _analyze_references(self):
        self.references = {}
        self.external_references = {}

        for type_name, type_def in self.types.items():
            for field_name, field_def in type_def['fields'].items():
                constraints = field_def.get('constraints', [])

                for constraint in constraints:
                    if isinstance(constraint, dict):
                        if 'references' in constraint:
                            ref = constraint['references']
                            is_external = any(c == 'external' or (isinstance(c, dict) and 'external' in c) for c in constraints)
                            target = (type_name, field_name, ref, False, is_external)

                            if field_def["type"].startswith("array<"):
                                print(f"In: {type_name}")
                                print(f"[ERROR] 'references' expects single target, not array!")
                                print(field_def)
                                exit(1)

                            if is_external:
                                self.external_references.setdefault(type_name, []).append(target)
                            else:
                                self.references.setdefault(type_name, []).append(target)

                        elif 'references_many' in constraint:
                            ref = constraint['references_many']
                            is_external = any(c == 'external' or (isinstance(c, dict) and 'external' in c) for c in constraints)
                            target = (type_name, field_name, ref, True, is_external)

                            if not field_def["type"].startswith("array<"):
                                print(f"In: {type_name}")
                                print(f"[ERROR] 'references_many' expects array, not single target!")
                                print(field_def)
                                exit(1)

                            if is_external:
                                self.external_references.setdefault(type_name, []).append(target)
                            else:
                                self.references.setdefault(type_name, []).append(target)

    def generate_all(self, output_dir: str = ".", output_dir_godot:str="."):
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        output_path_godot = Path(output_dir_godot)
        output_path_godot.mkdir(exist_ok=True)

        self.generate_gdscript_classes(output_path_godot / "generated_classes.gd")
        self.generate_data_loader(output_path_godot / "data_loader.gd")
        self.generate_excel_template(output_path / "data_template.xlsx")
        self.generate_excel_converter(output_path / "excel_to_json.py")
        self.generate_validator(output_path / "validate_data.py")

        print(f"Generated files in {output_path} and {output_path_godot}")

    def generate_gdscript_classes(self, output_path: Path):
        template = self.env.get_template('gdscript_classes.gd.j2')
        content = template.render(
            types=self.types,
            enums=self.enums,
            schema=self.schema
        )
        output_path.write_text(content, "utf-8")

    def generate_data_loader(self, output_path: Path):
        template = self.env.get_template('data_loader.gd.j2')
        content = template.render(
            types=self.types,
            references=self.references,
            external_references=self.external_references
        )
        output_path.write_text(content, "utf-8")

    def generate_excel_converter(self, output_path: Path):
        template = self.env.get_template('excel_to_json.py.j2')
        content = template.render(
            types=self.types
        )
        output_path.write_text(content, "utf-8")
        output_path.chmod(0o755)

    def generate_validator(self, output_path: Path):
        template = self.env.get_template('validator.py.j2')
        content = template.render(
            types=self.types,
            enums=self.enums,
            references=self.references,
            external_references=self.external_references,
            schema=self.schema
        )
        output_path.write_text(content, "utf-8")
        output_path.chmod(0o755)

    def generate_excel_template(self, output_path: Path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # pyright: ignore[reportArgumentType]

        excel_config = self.schema.get('excel', {})
        default_width = excel_config.get('column_width', {}).get('default', 20)

        for type_name in excel_config.get('sheet_order', self.types.keys()):
            type_def = self.types[type_name]
            ws = wb.create_sheet(type_name)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            col_idx = 1
            for field_name, field_def in type_def['fields'].items():
                cell = ws.cell(row=1, column=col_idx)
                cell.value = field_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

                desc = field_def.get('description', '')
                constraints = field_def.get('constraints', [])

                ref_info = []
                for c in constraints:
                    if isinstance(c, dict):
                        if 'references' in c:
                            ref_info.append(f"→ {c['references']}")
                        elif 'references_many' in c:
                            ref_info.append(f"→ {c['references_many']} (comma-separated)")

                if ref_info:
                    desc = f"{desc} | {' '.join(ref_info)}"

                desc_cell = ws.cell(row=2, column=col_idx)
                desc_cell.value = desc
                desc_cell.font = Font(italic=True, size=9)

                width = excel_config.get('column_width', {}).get(field_name, default_width)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

                col_idx += 1

            ws.freeze_panes = 'A3'

        wb.save(output_path)

    def _map_to_gdscript_type(self, field_def: Dict) -> str:
        type_str = field_def['type'] if isinstance(field_def, dict) else field_def

        if type_str == 'int':
            return 'int'
        elif type_str == 'float':
            return 'float'
        elif type_str == 'string':
            return 'String'
        elif type_str == 'bool':
            return 'bool'
        elif type_str.startswith('array<'):
            inner = type_str[6:-1]
            if inner in self.types:
                return f'Array[GeneratedDataClasses.{inner}]'
            inner_mapped = self._map_to_gdscript_type(inner)
            return f'Array[{inner_mapped}]'
        elif type_str in self.types:
            return f'GeneratedDataClasses.{type_str}'
        return 'Variant'

    def _get_gdscript_default(self, field_def: Dict) -> str:
        type_str = field_def['type']
        constraints = field_def.get('constraints', [])

        if 'optional' in constraints:
            return 'null'

        if type_str == 'int':
            return '0'
        elif type_str == 'float':
            return '0.0'
        elif type_str == 'string':
            return '""'
        elif type_str == 'bool':
            return 'false'
        elif type_str.startswith('array'):
            return '[]'
        elif type_str in self.types:
            return 'null'
        return 'null'

    def _map_to_python_type(self, field_def: Dict) -> str:
        type_str = field_def['type'] if isinstance(field_def, dict) else field_def

        if type_str == 'int':
            return 'int'
        elif type_str == 'float':
            return 'float'
        elif type_str == 'string':
            return 'str'
        elif type_str == 'bool':
            return 'bool'
        elif type_str.startswith('array<'):
            inner = type_str[6:-1]
            return f'List[{self._map_to_python_type(inner)}]'
        return 'Any'

    def _get_id_field(self, type_def: Dict) -> str:
        """Find the ID field for a type (looks for fields ending with 'id')"""
        for field_name in type_def['fields'].keys():
            if field_name.endswith('id'):
                return field_name
        return ""
        # TODO: return None here?

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python generate_from_schema_v2.py <schema.yaml> [output_dir]")
        sys.exit(1)

    schema_path = sys.argv[1]
    output_dir = sys.argv[3] if len(sys.argv) > 3 else "."
    output_dir_godot = sys.argv[2] if len(sys.argv) > 2 else "."

    generator = SchemaGenerator(schema_path)
    generator.generate_all(output_dir, output_dir_godot)