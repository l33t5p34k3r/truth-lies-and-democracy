#!/usr/bin/env python3
import openpyxl
import json
from pathlib import Path
import sys

def format_value(value, field_type: str):
    if field_type.startswith("array<"):
        if isinstance(value, list):
            return ", ".join(str(item) for item in value)
        return str(value)
    elif field_type == "bool":
        return "true" if value else "false"
    else:
        return value

def json_to_excel(json_path: str, excel_path: str, output_path: str):
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    wb = openpyxl.load_workbook(excel_path)

    if "StoryGroup" in wb.sheetnames and "StoryGroup" in json_data:
        ws = wb["StoryGroup"]
        headers = [cell.value for cell in ws[1]]
        
        field_types = {
            "group_id": "int",
            "stories": "array<int>",
        }

        for row_idx, entry in enumerate(json_data["StoryGroup"], start=3):
            for col_idx, header in enumerate(headers, start=1):
                if header in entry and header in field_types:
                    value = entry[header]
                    formatted_value = format_value(value, field_types[header])
                    ws.cell(row=row_idx, column=col_idx, value=formatted_value)

    if "Story" in wb.sheetnames and "Story" in json_data:
        ws = wb["Story"]
        headers = [cell.value for cell in ws[1]]
        
        field_types = {
            "story_id": "int",
            "news_headline": "string",
            "news_content": "string",
            "news_fake": "bool",
        }

        for row_idx, entry in enumerate(json_data["Story"], start=3):
            for col_idx, header in enumerate(headers, start=1):
                if header in entry and header in field_types:
                    value = entry[header]
                    formatted_value = format_value(value, field_types[header])
                    ws.cell(row=row_idx, column=col_idx, value=formatted_value)

    if "MediaPostGroup" in wb.sheetnames and "MediaPostGroup" in json_data:
        ws = wb["MediaPostGroup"]
        headers = [cell.value for cell in ws[1]]
        
        field_types = {
            "group_id": "int",
            "story_posts": "array<int>",
        }

        for row_idx, entry in enumerate(json_data["MediaPostGroup"], start=3):
            for col_idx, header in enumerate(headers, start=1):
                if header in entry and header in field_types:
                    value = entry[header]
                    formatted_value = format_value(value, field_types[header])
                    ws.cell(row=row_idx, column=col_idx, value=formatted_value)

    if "StoryPosts" in wb.sheetnames and "StoryPosts" in json_data:
        ws = wb["StoryPosts"]
        headers = [cell.value for cell in ws[1]]
        
        field_types = {
            "story_id": "int",
            "posts": "array<int>",
        }

        for row_idx, entry in enumerate(json_data["StoryPosts"], start=3):
            for col_idx, header in enumerate(headers, start=1):
                if header in entry and header in field_types:
                    value = entry[header]
                    formatted_value = format_value(value, field_types[header])
                    ws.cell(row=row_idx, column=col_idx, value=formatted_value)

    if "SocialMediaPost" in wb.sheetnames and "SocialMediaPost" in json_data:
        ws = wb["SocialMediaPost"]
        headers = [cell.value for cell in ws[1]]
        
        field_types = {
            "post_id": "int",
            "user_name": "string",
            "content_text": "string",
        }

        for row_idx, entry in enumerate(json_data["SocialMediaPost"], start=3):
            for col_idx, header in enumerate(headers, start=1):
                if header in entry and header in field_types:
                    value = entry[header]
                    formatted_value = format_value(value, field_types[header])
                    ws.cell(row=row_idx, column=col_idx, value=formatted_value)

    wb.save(output_path)
    print(f"Populated {output_path} with data from {json_path}")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python json_to_excel.py <input.json> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    json_to_excel(sys.argv[1], sys.argv[2], sys.argv[3])