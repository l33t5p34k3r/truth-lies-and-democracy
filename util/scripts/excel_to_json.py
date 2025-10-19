#!/usr/bin/env python3
import openpyxl
import json
from pathlib import Path
import sys

def convert_value(value_str: str, target_type: str):
    """Convert a string value to the target type"""
    value_str = value_str.strip()

    if target_type == "int":
        try:
            return int(value_str)
        except ValueError:
            print(f"Warning: Cannot convert '{value_str}' to int, using 0")
            return 0
    elif target_type == "float":
        try:
            return float(value_str)
        except ValueError:
            print(f"Warning: Cannot convert '{value_str}' to float, using 0.0")
            return 0.0
    elif target_type == "bool":
        lower = value_str.lower()
        return lower in ["true", "1", "yes", "y"]
    else:
        # string or unknown type
        return value_str

def excel_to_json(excel_path: str, output_path: str):
    wb = openpyxl.load_workbook(excel_path)
    all_data = {}

    if "StoryGroup" in wb.sheetnames:
        ws = wb["StoryGroup"]
        headers = [cell.value for cell in ws[1]]
        storygroup_data = []

        # Build field type mapping
        field_types = {
            "group_id": "int",
            "stories": "array<int>",
        }

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                continue

            entry = {}
            for header, value in zip(headers, row):
                if value is not None and header in field_types:
                    field_type = field_types[header]

                    # Handle array types
                    if field_type.startswith("array<"):
                        # TODO: this is really stupid
                        inner_type = field_type[6:-1]
                        if isinstance(value, list):
                            # Already a list (shouldn't happen in Excel, but handle it)
                            if inner_type in ["int", "float", "bool"]:
                                entry[header] = [convert_value(str(item), inner_type) for item in value]
                            else:
                                entry[header] = value
                        else:
                            # Convert comma-separated string to list with proper types
                            value_str = str(value)
                            raw_list = [x.strip() for x in value_str.split(",") if x.strip()]

                            # Convert each element to the correct type
                            if inner_type == "int":
                                entry[header] = [convert_value(x, "int") for x in raw_list]
                            elif inner_type == "float":
                                entry[header] = [convert_value(x, "float") for x in raw_list]
                            elif inner_type == "bool":
                                entry[header] = [convert_value(x, "bool") for x in raw_list]
                            else:
                                # string or custom type
                                entry[header] = raw_list

                    # Handle basic types
                    elif field_type == "int":
                        entry[header] = convert_value(str(value), "int")
                    elif field_type == "float":
                        entry[header] = convert_value(str(value), "float")
                    elif field_type == "bool":
                        entry[header] = convert_value(str(value), "bool")

                    # All other types: convert to string
                    else:
                        entry[header] = str(value) if not isinstance(value, (list, dict)) else value


            if entry:
                storygroup_data.append(entry)

        all_data["StoryGroup"] = storygroup_data

    if "Story" in wb.sheetnames:
        ws = wb["Story"]
        headers = [cell.value for cell in ws[1]]
        story_data = []

        # Build field type mapping
        field_types = {
            "story_id": "int",
            "news_headline": "string",
            "news_content": "string",
            "news_fake": "bool",
        }

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                continue

            entry = {}
            for header, value in zip(headers, row):
                if value is not None and header in field_types:
                    field_type = field_types[header]

                    # Handle array types
                    if field_type.startswith("array<"):
                        # TODO: this is really stupid
                        inner_type = field_type[6:-1]
                        if isinstance(value, list):
                            # Already a list (shouldn't happen in Excel, but handle it)
                            if inner_type in ["int", "float", "bool"]:
                                entry[header] = [convert_value(str(item), inner_type) for item in value]
                            else:
                                entry[header] = value
                        else:
                            # Convert comma-separated string to list with proper types
                            value_str = str(value)
                            raw_list = [x.strip() for x in value_str.split(",") if x.strip()]

                            # Convert each element to the correct type
                            if inner_type == "int":
                                entry[header] = [convert_value(x, "int") for x in raw_list]
                            elif inner_type == "float":
                                entry[header] = [convert_value(x, "float") for x in raw_list]
                            elif inner_type == "bool":
                                entry[header] = [convert_value(x, "bool") for x in raw_list]
                            else:
                                # string or custom type
                                entry[header] = raw_list

                    # Handle basic types
                    elif field_type == "int":
                        entry[header] = convert_value(str(value), "int")
                    elif field_type == "float":
                        entry[header] = convert_value(str(value), "float")
                    elif field_type == "bool":
                        entry[header] = convert_value(str(value), "bool")

                    # All other types: convert to string
                    else:
                        entry[header] = str(value) if not isinstance(value, (list, dict)) else value


            if entry:
                story_data.append(entry)

        all_data["Story"] = story_data

    if "MediaPostGroup" in wb.sheetnames:
        ws = wb["MediaPostGroup"]
        headers = [cell.value for cell in ws[1]]
        mediapostgroup_data = []

        # Build field type mapping
        field_types = {
            "group_id": "int",
            "story_posts": "array<int>",
        }

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                continue

            entry = {}
            for header, value in zip(headers, row):
                if value is not None and header in field_types:
                    field_type = field_types[header]

                    # Handle array types
                    if field_type.startswith("array<"):
                        # TODO: this is really stupid
                        inner_type = field_type[6:-1]
                        if isinstance(value, list):
                            # Already a list (shouldn't happen in Excel, but handle it)
                            if inner_type in ["int", "float", "bool"]:
                                entry[header] = [convert_value(str(item), inner_type) for item in value]
                            else:
                                entry[header] = value
                        else:
                            # Convert comma-separated string to list with proper types
                            value_str = str(value)
                            raw_list = [x.strip() for x in value_str.split(",") if x.strip()]

                            # Convert each element to the correct type
                            if inner_type == "int":
                                entry[header] = [convert_value(x, "int") for x in raw_list]
                            elif inner_type == "float":
                                entry[header] = [convert_value(x, "float") for x in raw_list]
                            elif inner_type == "bool":
                                entry[header] = [convert_value(x, "bool") for x in raw_list]
                            else:
                                # string or custom type
                                entry[header] = raw_list

                    # Handle basic types
                    elif field_type == "int":
                        entry[header] = convert_value(str(value), "int")
                    elif field_type == "float":
                        entry[header] = convert_value(str(value), "float")
                    elif field_type == "bool":
                        entry[header] = convert_value(str(value), "bool")

                    # All other types: convert to string
                    else:
                        entry[header] = str(value) if not isinstance(value, (list, dict)) else value


            if entry:
                mediapostgroup_data.append(entry)

        all_data["MediaPostGroup"] = mediapostgroup_data

    if "StoryPosts" in wb.sheetnames:
        ws = wb["StoryPosts"]
        headers = [cell.value for cell in ws[1]]
        storyposts_data = []

        # Build field type mapping
        field_types = {
            "story_id": "int",
            "posts": "array<int>",
        }

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                continue

            entry = {}
            for header, value in zip(headers, row):
                if value is not None and header in field_types:
                    field_type = field_types[header]

                    # Handle array types
                    if field_type.startswith("array<"):
                        # TODO: this is really stupid
                        inner_type = field_type[6:-1]
                        if isinstance(value, list):
                            # Already a list (shouldn't happen in Excel, but handle it)
                            if inner_type in ["int", "float", "bool"]:
                                entry[header] = [convert_value(str(item), inner_type) for item in value]
                            else:
                                entry[header] = value
                        else:
                            # Convert comma-separated string to list with proper types
                            value_str = str(value)
                            raw_list = [x.strip() for x in value_str.split(",") if x.strip()]

                            # Convert each element to the correct type
                            if inner_type == "int":
                                entry[header] = [convert_value(x, "int") for x in raw_list]
                            elif inner_type == "float":
                                entry[header] = [convert_value(x, "float") for x in raw_list]
                            elif inner_type == "bool":
                                entry[header] = [convert_value(x, "bool") for x in raw_list]
                            else:
                                # string or custom type
                                entry[header] = raw_list

                    # Handle basic types
                    elif field_type == "int":
                        entry[header] = convert_value(str(value), "int")
                    elif field_type == "float":
                        entry[header] = convert_value(str(value), "float")
                    elif field_type == "bool":
                        entry[header] = convert_value(str(value), "bool")

                    # All other types: convert to string
                    else:
                        entry[header] = str(value) if not isinstance(value, (list, dict)) else value


            if entry:
                storyposts_data.append(entry)

        all_data["StoryPosts"] = storyposts_data

    if "SocialMediaPost" in wb.sheetnames:
        ws = wb["SocialMediaPost"]
        headers = [cell.value for cell in ws[1]]
        socialmediapost_data = []

        # Build field type mapping
        field_types = {
            "post_id": "int",
            "user_name": "string",
            "content_text": "string",
        }

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                continue

            entry = {}
            for header, value in zip(headers, row):
                if value is not None and header in field_types:
                    field_type = field_types[header]

                    # Handle array types
                    if field_type.startswith("array<"):
                        # TODO: this is really stupid
                        inner_type = field_type[6:-1]
                        if isinstance(value, list):
                            # Already a list (shouldn't happen in Excel, but handle it)
                            if inner_type in ["int", "float", "bool"]:
                                entry[header] = [convert_value(str(item), inner_type) for item in value]
                            else:
                                entry[header] = value
                        else:
                            # Convert comma-separated string to list with proper types
                            value_str = str(value)
                            raw_list = [x.strip() for x in value_str.split(",") if x.strip()]

                            # Convert each element to the correct type
                            if inner_type == "int":
                                entry[header] = [convert_value(x, "int") for x in raw_list]
                            elif inner_type == "float":
                                entry[header] = [convert_value(x, "float") for x in raw_list]
                            elif inner_type == "bool":
                                entry[header] = [convert_value(x, "bool") for x in raw_list]
                            else:
                                # string or custom type
                                entry[header] = raw_list

                    # Handle basic types
                    elif field_type == "int":
                        entry[header] = convert_value(str(value), "int")
                    elif field_type == "float":
                        entry[header] = convert_value(str(value), "float")
                    elif field_type == "bool":
                        entry[header] = convert_value(str(value), "bool")

                    # All other types: convert to string
                    else:
                        entry[header] = str(value) if not isinstance(value, (list, dict)) else value


            if entry:
                socialmediapost_data.append(entry)

        all_data["SocialMediaPost"] = socialmediapost_data

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, indent=2, ensure_ascii=False)

    print(f"Converted {excel_path} to {output_path}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python excel_to_json.py <input.xlsx> <output.json>")
        sys.exit(1)

    excel_to_json(sys.argv[1], sys.argv[2])